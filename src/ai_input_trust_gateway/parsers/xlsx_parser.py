"""xlsx parser: openpyxl → sheet/row/cell model (hidden cells included)."""

from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

from ai_input_trust_gateway.core.document import ParsedDocument, Sheet, TextParagraph, TextRun
from ai_input_trust_gateway.parsers.base import register_parser
from ai_input_trust_gateway.parsers.ooxml_raw import OOXMLRawReader


def _normalize_color(val: str | None) -> str | None:
    if not val:
        return None
    v = val.strip().lstrip("#").upper()
    if len(v) == 6 and all(c in "0123456789ABCDEF" for c in v):
        return v
    return None


@register_parser("xlsx", ("xlsx", "xlsm"), sniff=None)
class XlsxParser:
    """Parse .xlsx into :class:`ParsedDocument`."""

    def parse(self, path: Path) -> ParsedDocument:
        doc_model = ParsedDocument(kind="xlsx", source_path=str(path))
        wb = load_workbook(str(path), data_only=False)

        for ws in wb.worksheets:
            sheet = Sheet(
                name=ws.title,
                state=str(ws.sheet_state),  # visible | hidden | veryHidden
                merged_ranges=[str(r) for r in ws.merged_cells.ranges],
            )
            # hidden rows / cols (only record non-empty hidden regions in detector)
            for row_idx, dim in ws.row_dimensions.items():
                if dim.hidden:
                    sheet.rows_hidden.append(row_idx)
            for col_idx, dim in ws.column_dimensions.items():
                if dim.hidden:
                    sheet.cols_hidden.append(col_idx)
            doc_model.sheets.append(sheet)

            # cells → paragraphs (row number = paragraph.index)
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value is None:
                        continue
                    text = str(cell.value)
                    if not text:
                        continue
                    color = None
                    size = None
                    try:
                        font = cell.font
                        if font and font.color and font.color.rgb:
                            color = _normalize_color(str(font.color.rgb))
                        if font and font.size:
                            size = float(font.size)
                    except Exception:  # noqa: BLE001 — style access can fail
                        pass
                    is_hidden = False
                    # hidden cell heuristic: row or col is hidden
                    if cell.row in sheet.rows_hidden or cell.column in sheet.cols_hidden:
                        is_hidden = True
                    run = TextRun(text=text, font_size=size, color=color, is_hidden=is_hidden)
                    doc_model.paragraphs.append(
                        TextParagraph(
                            text=text,
                            index=cell.row,
                            runs=[run],
                            raw={"sheet": ws.title, "row": cell.row, "col": cell.column},
                        )
                    )

        # core properties
        try:
            props = wb.properties
            doc_model.metadata = {
                "creator": props.creator,
                "title": props.title,
                "subject": props.subject,
                "keywords": props.keywords,
                "description": props.description,
            }
        except Exception:  # noqa: BLE001
            doc_model.metadata = {}

        # raw OOXML parts
        try:
            with OOXMLRawReader(path) as reader:
                doc_model.ooxml_parts = {name: reader.part_bytes(name) for name in reader.part_names()}
        except Exception as exc:  # noqa: BLE001
            doc_model.warnings.append(f"failed to read raw OOXML parts: {exc}")

        return doc_model
